import openpyxl
import os
import jieba.posseg as pseg
from utils import fun_tongji,sim_tool


yanbaodir='E:\\桌面\\期货\\研报ocr'
jiaoyishoucedir='E:\\桌面\\期货\\期货OCR'
sim_path='E:\\桌面\\期货\\研报ocr\\相似度计算结果.xlsx'


#处理种子词，形成短语
def dealseed():
    file = fun_tongji.Excel(sim_path)
    seedwb=file.getwb('筛选结果（研报）')
    cizuwb=file.createsheet('短语发现（研报）')
    seed=[]
    for i in range(1,seedwb.max_row+1):
        if seedwb.cell(i,1).value not in ['None',''] and seedwb.cell(i,1).value!=None:
            seed.append(seedwb.cell(i,1).value)
    seed=list(set(seed))
    print(seed)
    cizu=[]
    sim=sim_tool.wordmine()
    textlist=sim.getalltext(yanbaodir)
    print(len(textlist))

    allvocab=[]
    allvocabflag=[]
    for text in textlist:
        #print(text)

        segment=pseg.cut(text)
        vocab=[]
        vocabflag=[]
        for word,flag in segment:
            vocab.append(word)
            vocabflag.append(word+flag)
        allvocab.append(vocab)
        allvocabflag.append(vocabflag)
    for oneword in seed:
        wordcizu=list(set(sim.findcizu(allvocab,allvocabflag,oneword)))
        print(wordcizu)
        if wordcizu==[]:
            print(oneword)
        cizu+=wordcizu
    cizu=list(set(cizu))
    k=1
    for onecizu in cizu:
        cizuwb.cell(row=k,column=1,value=onecizu)
        k+=1
    file.save(sim_path)

# def bijiao():
#     file = openpyxl.load_workbook('E:\\桌面\\期货\\研报ocr\\相似度计算结果.xlsx')
#     wb1=file['短语发现（交易手册）']
#     wordlist1=[]
#     for i in range(1,wb1.max_row+1):
#         wordlist1.append(wb1.cell(i,1).value)
#     print(wordlist1)
#     wb2 = file['短语发现（交易手册）1']
#     wordlist2 = []
#     for i in range(1, wb2.max_row + 1):
#         wordlist2.append(wb2.cell(i, 1).value)
#     print(wordlist2)
#     for word in wordlist1:
#         if word not in wordlist2:
#             print(word)

dealseed()