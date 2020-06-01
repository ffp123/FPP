import numpy as np
import openpyxl
from utils import fun_tongji,sim_tool
import jieba
import jieba.posseg as pseg
import Levenshtein
embedding_path='E:\\桌面\\期货\\bert_embedding.txt'
reci_path='E:\\桌面\\期货\\热词.xlsx'
result_path='E:\\桌面\\期货\\研报ocr\\相似度计算结果.xlsx'

stop = open('E:\\KGlearning\\finance-futures\\FPP3-m\\baidu_stopwords.txt', 'r+', encoding='utf-8')
stopword = stop.read().split("\n")


#获取热词的词向量
def get_reci_vec():
    sim=sim_tool.similarity()
    vocab,vecdict=sim.load_wordvector(embedding_path)
    reciexcel=fun_tongji.Excel(reci_path)
    recisheets=reciexcel.getSheetName()
    recilist=[]
    for ws in recisheets:
        wb=reciexcel.getwb(ws)
        for i in range(2,wb.max_row+1):
            reci=wb.cell(i,2).value
            if reci not in ['None',''] and reci !=None:
                recilist.append(reci)

    recivec_dict=sim.get_reci_vec(vocab,vecdict,recilist)
    print(recivec_dict)
    return recivec_dict,vocab,vecdict

# def xx():
#     reciexcel=openpyxl.load_workbook(reci_path)
#     resultexcel=openpyxl.load_workbook(result_path)
#     resultwb=resultexcel['热词']
#     recisheets=reciexcel.sheetnames
#     j=2
#     for ws in recisheets:
#         wb=reciexcel[ws]
#         for i in range(2,wb.max_row+1):
#             reci=wb.cell(i,2).value
#             if reci not in ['None',''] and reci !=None:
#                 reci_seg = pseg.cut(reci)
#                 reci_list=[]
#                 for word,flag in reci_seg:
#                     if flag in ['n', 'ng', 'nr', 'nrfg', 'nrt', 'ns', 'nt', 'nz', 's', 't', 'tg',
#                                 'vn'] and word not in stopword:
#                         reci_list.append(word)
#                 resultwb.cell(row=j,column=1,value=reci)
#                 if len(reci_list)==0:
#                     resultwb.cell(row=j, column=2, value=reci)
#                 else:
#                     resultwb.cell(row=j,column=2,value=''.join(reci_list))
#                 j+=1
#     resultexcel.save(result_path)

#计算相似度测试函数
def calsim():
    file=fun_tongji.Excel(result_path)
    reciwb=file.getwb('热词')
    wb=file.getwb('短语发现（交易手册）')
    resultwb=file.createsheet('相似度比较（交易手册短语）')
    recivec_dict,vocab,vector_dic=get_reci_vec()
    sim=sim_tool.similarity()
    resultwb.cell(row=1, column=1, value='待比较词')
    resultwb.cell(row=1, column=2, value='热词')
    resultwb.cell(row=1, column=3, value='余弦相似度')
    resultwb.cell(row=1, column=4, value='jaro-winkler相似度')
    resultwb.cell(row=1, column=5, value='jaccard相似度')
    k=2
    for i in range(2,wb.max_row+1):
        print(i)
        key=wb.cell(i,1).value
        if key not in ['None',''] and key !=None:
             label = fun_tongji.get_alpha_str(key)
             word = key.replace(label, '')
             word_vec=sim.get_word_vec(word)
             for j in range(2, reciwb.max_row + 1):
                reci = reciwb.cell(j,1).value
                cos=sim.cossim(word_vec,recivec_dict[reci])*0.5+0.5
                jarov=Levenshtein.jaro_winkler(reci,word)
                jacv=sim.jaccard(reci,word)
                resultwb.cell(row=k,column=1,value=word)
                resultwb.cell(row=k, column=2, value=reci)
                resultwb.cell(row=k,column=3,value=cos)
                resultwb.cell(row=k, column=4, value=jarov)
                resultwb.cell(row=k, column=5, value=jacv)
                k+=1
    file.save(result_path)

#筛选计算相似度的结果
def shaixuan():
    file=fun_tongji.Excel(result_path)
    wb=file.getwb('相似度比较（研报短语）')
    resultwb=file.createsheet('筛选结果（研报短语）')
    k=1
    for i in range(2,wb.max_row+1):
        if wb.cell(i, 3).value not in ['','None'] and wb.cell(i, 3).value!=None:
            cosvalue=wb.cell(i,3).value
            jarov=wb.cell(i,4).value
            jacv=wb.cell(i,5).value
            if cosvalue>=0.80 or jarov>=0.80 or jacv>=0.70:
                resultwb.cell(row=k,column=1,value=wb.cell(i,1).value)
                resultwb.cell(row=k, column=2, value=wb.cell(i, 2).value)
                resultwb.cell(row=k, column=3, value=cosvalue)
                resultwb.cell(row=k, column=4, value=jarov)
                resultwb.cell(row=k, column=5, value=jacv)
                k+=1
    file.save(result_path)

def quchong():
    file = fun_tongji.Excel(result_path)
    wb = file.getwb('筛选结果（交易手册短语）')
    wordlist=[]
    for i in range(1,wb.max_row+1):
        if wb.cell(i,1).value not in wordlist:
            wordlist.append(wb.cell(i,1).value)
    k=1
    for word in wordlist:
        wb.cell(row=k,column=8,value=word)
        k+=1
    file.save(result_path)

