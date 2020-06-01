import openpyxl
import ast
import os
from utils import fun_tongji
cip_path='E:\\桌面\\期货\\研报ocr\\词频统计.xlsx'
ny_path='E:\\桌面\\期货\\研报ocr\\能源期货.xlsx'
ncp_path='E:\\桌面\\期货\\研报ocr\\农产品期货.xlsx'
js_path='E:\\桌面\\期货\\研报ocr\\金属期货.xlsx'

#统计每个期货词频
def tongjichanpin():
    file=fun_tongji.Excel(cip_path)
    file1 = fun_tongji.Excel(ny_path)
    resultwb=file.createsheet('能源类期货统计结果')
    huizong = file.getwb('各类期货统计结果(汇总)')
    worksheets = file1.getSheetName()
    tongji=fun_tongji.tongji()
    allsortdict=tongji.tongjichanpin(worksheets,file1,resultwb)

    huizong.cell(row=1, column=7, value='能源类期货汇总')
    huizong.cell(row=1, column=8, value='词频统计')
    i=2
    for key, value in allsortdict.items():
        huizong.cell(row=i, column=7, value=key)
        huizong.cell(row=i, column=8, value=value)
        i += 1
    file.save(cip_path)

#统计所有期货词频
def tongjiall():
    path = 'E:\\桌面\\期货\\研报ocr'
    resultfile = fun_tongji.Excel(cip_path)
    wb2 = resultfile.getwb('各类期货统计结果(汇总)')
    dirs = os.listdir(path)
    tongji=fun_tongji.tongji()
    sortdict=tongji.tonngjiall(dirs,path)
    wb2.cell(row=1, column=1, value='所有期货统计结果')
    wb2.cell(row=1, column=2, value='词频统计')
    i=2
    for key,value in sortdict.items():
        wb2.cell(row=i, column=1, value=key)
        wb2.cell(row=i, column=2, value=value)
        i+=1
    resultfile.save(cip_path)

#统计各类期货专有词与共有词
def tongjicom_and_spec():
    resultfile = fun_tongji.Excel(cip_path)
    ws=resultfile.getwb('各类期货统计结果(汇总)')
    huafen=resultfile.createsheet('不同类别期货共有词专有词统计(20200503)')
    tongji=fun_tongji.tongji()
    sortjj2dict, sortncpjj2dict, sortnyjj2dict, sortjsjj2dict,ncpdict,nydict,jsdict=tongji.tongjicom_and_spec(ws)
    i=2
    for key,value in sortjj2dict.items():
        huafen.cell(row=i,column=1,value=key)
        huafen.cell(row=i, column=2, value=nydict[key])
        huafen.cell(row=i, column=3, value=ncpdict[key])
        huafen.cell(row=i, column=4, value=jsdict[key])
        i+=1
    i = 2
    for key,value in sortncpjj2dict.items():
        huafen.cell(row=i, column=7, value=key)
        huafen.cell(row=i, column=8, value=value)
        i+=1
    i = 2
    for key,value in sortnyjj2dict.items():
        huafen.cell(row=i, column=5, value=key)
        huafen.cell(row=i, column=6, value=value)
        i+=1
    i = 2
    for key, value in sortjsjj2dict.items():
        huafen.cell(row=i, column=9, value=key)
        huafen.cell(row=i, column=10, value=value)
        i += 1
    resultfile.save(cip_path)


#统计每种期货专有词
def shaixuanchanp():
    file=fun_tongji.Excel(cip_path)
    prowb=file.getwb('农产品类期货统计结果')
    resultwb=file.createsheet('农产品筛选')
    tongji=fun_tongji.tongji()
    tongji.shaixuanchanp(resultwb,prowb)

    file.save(cip_path)

#找出各类别期货下同时出现在每种期货中的词
def findcommon_in_label():
    file = fun_tongji.Excel('E:\\桌面\\期货\\期货词频统计\\期货词频统计.xlsx')
    filews=file.getwb('能源筛选')
    prodict={}
    prodictvocab={}
    for j in range(3, filews.max_column + 1, 2):
        prodict[filews.cell(1,j).value]={}
        prodictvocab[filews.cell(1, j).value] = []
        for i in range(2,filews.max_row+1):
            if filews.cell(i,j).value!='' and filews.cell(i,j).value!=None:
                prodict[filews.cell(1,j).value][filews.cell(i,j).value]=filews.cell(i,j+1).value
                prodictvocab[filews.cell(1, j).value].append(filews.cell(i,j).value)
    base=prodictvocab['LLDPE期货词(专有)']
    for key,value in prodictvocab.items():
        jiaoji=list(set(base).intersection(set(value)))
        base=jiaoji
    print(base)


#获取各类期货下的共有词
def extract():
    file=fun_tongji.Excel(cip_path)
    allwordws=file.getwb('不同类别期货共有词专有词统计(20200503)')
    productws=file.getwb('金属筛选')
    resultws=file.createsheet('金属期货相关词（20200503）')
    tongji=fun_tongji.tongji()
    word_score_dict,productword=tongji.extract(allwordws,productws)
    resultws.cell(row=1,column=1,value='金属期货共有词')
    resultws.cell(row=1, column=2, value='置信度')
    i=2
    for key,value in word_score_dict.items():
        resultws.cell(row=i,column=1,value=key)
        resultws.cell(row=i, column=2, value=value)
        i+=1
    j=3
    for prod,worddict in productword.items():
        resultws.cell(row=1, column=j, value=prod)
        resultws.cell(row=1, column=j+1, value='词频统计')
        i=2
        for word,freq in worddict.items():
            #if word not in word_in_common:
            resultws.cell(row=i, column=j, value=word)
            resultws.cell(row=i, column=j + 1, value=freq)
            i+=1

        j+=2
    file.save(cip_path)

#按词性分类
def fenlei():
    file = fun_tongji.Excel(cip_path)
    ws=file.getwb('金属期货相关词（20200503）')
    resultws=file.createsheet('金属期货相关词(按词性)')

    tongji=fun_tongji.tongji()
    tongji.fenlei(ws,resultws)

    file.save(cip_path)

