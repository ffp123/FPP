import jieba
import jieba.analyse
import jieba.posseg as pseg
import openpyxl
import ast
import os

def baikezhishijiagong():
    p = open(r'玉米.txt', 'r', encoding = 'utf-8')
    q = open(r'玉米识别结果.txt', 'w', encoding = 'utf-8')
    r = open(r'玉米识别结果2.txt', 'w', encoding = 'utf-8')
    resultlist=[]
    for line in p.readlines():
        words = pseg.cut(line)
        for word, flag in words:
            if flag in ['n','ng','nr','nrfg','nrt','ns','nt','nz','s','t','tg']:
                r.write(str(word) + str(flag) + '\t')
                resultlist.append(str(word) + str(flag))
        r.write('\n')
    resultlist=list(set(resultlist))
    for item in resultlist:
        q.write(item+'\t')
    p.close()
    q.close()
    r.close()

def jiaoyishouce():
    file1=openpyxl.load_workbook('E:\\桌面\\期货\\能源期货.xlsx')
    worksheets=file1.sheetnames
    print(worksheets)
    for ws in worksheets:
        if ws!='交易手册划分汇总':
            wb2=file1[ws]
            wb2.cell(row=1, column=1, value='相关因素')
            wb2.cell(row=1, column=2, value='语料原文')
            wb2.cell(row=1, column=3, value='分词结果')
            wb2.cell(row=1, column=4, value='名词识别')
            wb2.cell(row=1, column=5, value='关键词提取（TF-IDF）')
            wb2.cell(row=1, column=6, value='关键词提取（textrank）')
            wb2.cell(row=1, column=7, value='相关因素(分词)')

            for i in range(2,wb2.max_row+1):
                factor=wb2.cell(i,1).value
                fac_list2 = pseg.cut(factor)
                facresult2 = []
                for word, flag in fac_list2:
                    # if flag not in ['w','wkz','wky','wyz','wyy','wj','ww','wt','wd','wf','wm','wn','ws','wp','wb','wh']:
                    if flag != 'x':
                        facresult2.append(str(word))
                wb2.cell(row=i, column=7, value=str(facresult2))
                context=wb2.cell(i,2).value
                # seg_list=jieba.cut(context)
                # segresult=[]
                # for word in seg_list:
                #     segresult.append(word)

                seg_list2 = pseg.cut(context)
                segresult2 = []
                for word,flag in seg_list2:
                    #if flag not in ['w','wkz','wky','wyz','wyy','wj','ww','wt','wd','wf','wm','wn','ws','wp','wb','wh']:
                    if flag!='x':
                        segresult2.append(str(word))
                #wb2.cell(row=i+1,column=1,value=wb1.cell(1,i).value)
                #wb2.cell(row=i, column=2, value=str(segresult))
                wb2.cell(row=i, column=3, value=str(segresult2))
                entity_list=[]
                words=pseg.cut(context)
                for word, flag in words:
                    #print(word,flag)
                    if flag in ['n','ng','nr','nrfg','nrt','ns','nt','nz','s','t','tg','vn']:
                        entity_list.append(str(word) + str(flag))
                #entity_list=list(set(entity_list))
                entity_freq_list={}
                for item in entity_list:
                    if item not in entity_freq_list:
                        entity_freq_list[item]=1
                    else:
                        entity_freq_list[item]+=1
                wb2.cell(row=i, column=4, value=str(entity_freq_list))
                keywords1 = jieba.analyse.extract_tags(context, topK=30, withWeight=True, allowPOS=('n','ng','nr','nrfg','nrt','ns','nt','nz','s','t','tg','vn'))
                keywordlist1={}
                for item in keywords1:
                    keywordlist1[item[0]]=item[1]
                wb2.cell(row=i, column=5, value=str(keywordlist1))
                keywords2 = jieba.analyse.textrank(context, topK=30, withWeight=True, allowPOS=(
                'n', 'ng', 'nr', 'nrfg', 'nrt', 'ns', 'nt', 'nz', 's', 't', 'tg', 'vn'))
                keywordlist2 = {}
                for item in keywords2:
                    keywordlist2[item[0]]=item[1]
                wb2.cell(row=i, column=6, value=str(keywordlist2))

    file1.save('E:\\桌面\\期货\\能源期货.xlsx')

def tongji():
    file1 = openpyxl.load_workbook('E:\\桌面\\期货\\期货词频统计\\农产品期货.xlsx')
    resultfile=openpyxl.load_workbook('E:\\桌面\\期货\\期货词频统计\\期货词频统计.xlsx')
    resultwb=resultfile.create_sheet('农产品类期货统计结果')
    huizong=resultfile.create_sheet('农产品类期货统计结果(汇总)')
    rowrecord=1
    worksheets = file1.sheetnames
    print(worksheets)
    allvocab={}
    for ws in worksheets:
        if ws != '交易手册划分汇总':
            wb2 = file1[ws]
            itemvocab={}
            print(wb2.max_row)
            for i in range(2, wb2.max_row + 1):
                if wb2.cell(i,1).value!=None and wb2.cell(i,1).value!='':
                    worddict=ast.literal_eval(wb2.cell(i,4).value)
                    # print(worddict)
                    # print(type(worddict))
                    for k,v in worddict.items():
                        if k not in itemvocab:
                            itemvocab[k]=v
                        else:
                            itemvocab[k]+=v
                        if k not in allvocab:
                            allvocab[k]=v
                        else:
                            allvocab[k]+=v
            resultwb.cell(row=rowrecord,column=1,value=ws)
            sortdict=sorted(itemvocab.items(), key=lambda item: item[1], reverse=True)
            resultwb.cell(row=rowrecord, column=2, value=str(sortdict))
            rowrecord+=1
    #file=open('金属期货汇总.txt','w',encoding='utf-8')
    sortall=sorted(allvocab.items(), key=lambda item: item[1], reverse=True)
    sortdict = {}
    for item in sortall:
        print(item[0], item[1])
        sortdict[item[0]] = item[1]
    i = 1
    for key, value in sortdict.items():
        huizong.cell(row=i, column=1, value=key)
        huizong.cell(row=i, column=2, value=value)
        i += 1
    # resultwb.cell(row=rowrecord, column=1, value='汇总结果')
    # resultwb.cell(row=rowrecord, column=2, value=str(sortall))
    # for one in sortall:
    #     file.write(str(one)+'\n')
    resultfile.save('E:\\桌面\\期货\\期货词频统计\\期货词频统计.xlsx')

def jieguobijiao():
    compflie=openpyxl.load_workbook('结果比较.xlsx')
    ws=compflie['手工标注']
    for i in range(2,ws.max_row+1):
        text=ws.cell(i,4).value
        segtext=[]
        flagtext=[]
        entitylist=[]
        otherentity=[]
        segresult = pseg.cut(text)
        for word,flag in segresult:
            segtext.append(word)
            flagtext.append(str(word) + str(flag))
            if flag in ['n', 'ng', 'nr', 'nrfg', 'nrt', 'ns', 'nt', 's', 't', 'tg', 'vn']:
                entitylist.append(str(word) + str(flag))
            if flag=='nz':
                otherentity.append(str(word) + str(flag))
        ws.cell(row=i, column=5, value=str(segtext))
        ws.cell(row=i, column=6, value=str(flagtext))
        ws.cell(row=i, column=7, value=str(entitylist))
        ws.cell(row=i, column=8, value=str(otherentity))

    compflie.save('结果比较.xlsx')

def tongjiall():
    path = 'E:\\桌面\\期货\\期货词频统计'
    resultfile = openpyxl.load_workbook('E:\\桌面\\期货\\期货词频统计\\期货词频统计.xlsx')
    wb2 = resultfile.create_sheet('所有期货统计结果')
    dirs = os.listdir(path)
    allvocab={}
    for dir in dirs:
        if dir!='期货词频统计.xlsx':
            file=openpyxl.load_workbook(path+'\\'+dir)
            worksheets = file.sheetnames
            for ws in worksheets:
                if ws != '交易手册划分汇总':
                    wb=file[ws]
                    for i in range(2, wb.max_row + 1):
                        if wb.cell(i, 1).value != None and wb.cell(i, 1).value != '':
                            worddict = ast.literal_eval(wb.cell(i, 4).value)
                            # print(worddict)
                            # print(type(worddict))
                            for k, v in worddict.items():
                                if k not in allvocab:
                                    allvocab[k] = v
                                else:
                                    allvocab[k] += v

    sortvocab= sorted(allvocab.items(), key=lambda item: item[1], reverse=True)
    print(sortvocab)
    sortdict={}
    for item in sortvocab:
        print(item[0],item[1])
        sortdict[item[0]]=item[1]
    i=1
    for key,value in sortdict.items():
        wb2.cell(row=i, column=1, value=key)
        wb2.cell(row=i, column=2, value=value)
        i+=1
    resultfile.save('E:\\桌面\\期货\\期货词频统计\\期货词频统计.xlsx')
    # file2 = open('所有期货汇总.txt', 'w', encoding='utf-8')
    # for one in sortdict:
    #     file2.write(str(one)+'\n')

def tongjiall2():
    resultfile = openpyxl.load_workbook('E:\\桌面\\期货\\期货词频统计\\期货词频统计.xlsx')
    nongchanpin=resultfile['农产品类期货统计结果(汇总)']
    jinshu=resultfile['金属类期货统计结果(汇总)']
    nengyuan = resultfile['能源类期货统计结果(汇总)']
    huafen=resultfile.create_sheet('不同类别期货相关词统计')
    ncpdict={}
    ncpvocab=[]
    for i in range(1,nongchanpin.max_row+1):
        word=nongchanpin.cell(i,1).value
        cipin=int(nongchanpin.cell(i,2).value)
        ncpdict[word]=cipin
        ncpvocab.append(word)
    jsdict={}
    jsvocab=[]
    for i in range(1,jinshu.max_row+1):
        word=jinshu.cell(i,1).value
        cipin=int(jinshu.cell(i,2).value)
        jsdict[word]=cipin
        jsvocab.append(word)
    nydict={}
    nyvocab=[]
    for i in range(1,nengyuan.max_row+1):
        word=nengyuan.cell(i,1).value
        cipin=int(nengyuan.cell(i,2).value)
        nydict[word]=cipin
        nyvocab.append(word)
    jj1=list(set(ncpvocab).intersection(set(nyvocab)))
    jj2 = list(set(jj1).intersection(set(jsvocab)))
    ncp_jj2=list(set(ncpvocab)-set(jj2))
    ny_jj2=list(set(nyvocab)-set(jj2))
    js_jj2 = list(set(jsvocab) - set(jj2))
    jj2dict={}
    for item in jj2:
        jj2dict[item]=nydict[item]
    sortjj2 = sorted(jj2dict.items(), key=lambda item: item[1], reverse=True)
    sortjj2dict={}
    for item in sortjj2:
        #print(item[0],item[1])
        sortjj2dict[item[0]]=item[1]
    i=2
    for key,value in sortjj2dict.items():
        huafen.cell(row=i,column=1,value=key)
        huafen.cell(row=i, column=2, value=nydict[key])
        huafen.cell(row=i, column=3, value=ncpdict[key])
        huafen.cell(row=i, column=4, value=jsdict[key])
        i+=1
    ncpjj2dict = {}
    for item in ncp_jj2:
        ncpjj2dict[item] = ncpdict[item]
    sortncpjj2 = sorted(ncpjj2dict.items(), key=lambda item: item[1], reverse=True)
    sortncpjj2dict = {}
    for item in sortncpjj2:
        # print(item[0],item[1])
        sortncpjj2dict[item[0]] = item[1]
    i = 2
    for key,value in sortncpjj2dict.items():
        huafen.cell(row=i, column=7, value=key)
        huafen.cell(row=i, column=8, value=value)
        i+=1
    nyjj2dict = {}
    for item in ny_jj2:
        nyjj2dict[item] = nydict[item]
    sortnyjj2 = sorted(nyjj2dict.items(), key=lambda item: item[1], reverse=True)
    sortnyjj2dict = {}
    for item in sortnyjj2:
        # print(item[0],item[1])
        sortnyjj2dict[item[0]] = item[1]
    i = 2
    for key,value in sortnyjj2dict.items():
        huafen.cell(row=i, column=5, value=key)
        huafen.cell(row=i, column=6, value=value)
        i+=1
    jsjj2dict = {}
    for item in js_jj2:
        jsjj2dict[item] = jsdict[item]
    sortjsjj2 = sorted(jsjj2dict.items(), key=lambda item: item[1], reverse=True)
    sortjsjj2dict = {}
    for item in sortjsjj2:
        # print(item[0],item[1])
        sortjsjj2dict[item[0]] = item[1]
    i = 2
    for key, value in sortjsjj2dict.items():
        huafen.cell(row=i, column=9, value=key)
        huafen.cell(row=i, column=10, value=value)
        i += 1

    resultfile.save('E:\\桌面\\期货\\期货词频统计\\期货词频统计.xlsx')

def tongjichanpin():
    file=openpyxl.load_workbook('E:\\桌面\\期货\\期货词频统计\\期货词频统计.xlsx')
    file1 = openpyxl.load_workbook('E:\\桌面\\期货\\期货词频统计\\能源期货.xlsx')
    resultwb=file.create_sheet('能源类期货统计结果')
    worksheets = file1.sheetnames
    print(worksheets)
    j=1
    for ws in worksheets:
        if ws != '交易手册划分汇总':
            wb2 = file1[ws]
            itemvocab = {}

            for i in range(2, wb2.max_row + 1):
                if wb2.cell(i, 1).value != None and wb2.cell(i, 1).value != '':
                    worddict = ast.literal_eval(wb2.cell(i, 4).value)
                    # print(worddict)
                    # print(type(worddict))
                    for k, v in worddict.items():
                        if k not in itemvocab:
                            itemvocab[k] = v
                        else:
                            itemvocab[k] += v

            sortvocab = sorted(itemvocab.items(), key=lambda item: item[1], reverse=True)
            sortdict={}
            for item in sortvocab:
                print(item[0], item[1])
                sortdict[item[0]] = item[1]
            resultwb.cell(row=1, column=j, value=str(ws+'期货词'))
            resultwb.cell(row=1, column=j+1, value='词频统计')
            n=2
            for key,value in sortdict.items():
                resultwb.cell(row=n, column=j, value=key)
                resultwb.cell(row=n, column=j + 1, value=value)
                n+=1
            j+=2

    file.save('E:\\桌面\\期货\\期货词频统计\\期货词频统计.xlsx')

def shaixuanchanp():
    file=openpyxl.load_workbook('E:\\桌面\\期货\\期货词频统计\\期货词频统计.xlsx')
    prowb=file['能源类期货统计结果']
    resultwb=file['能源筛选']
    commonvocab=[]
    for i in range(2,resultwb.max_row+1):
        if resultwb.cell(i,1).value!=None and resultwb.cell(i,1).value!='':
            commonvocab.append(resultwb.cell(i,1).value)

    for j in range(1,prowb.max_column+1,2):
        resultwb.cell(row=1,column=j+2,value=str(prowb.cell(1,j).value+'(专有)'))
        resultwb.cell(row=1, column=j + 3, value=prowb.cell(1, j+1).value)
        prodict={}
        provocab=[]
        for i in range(2,prowb.max_row+1):
            if prowb.cell(i,j).value!=None and prowb.cell(i,j).value!='':
                provocab.append(prowb.cell(i,j).value)
                prodict[prowb.cell(i,j).value]=prowb.cell(i,j+1).value
        n=2
        for key,value in prodict.items():
            if key in provocab and key not in commonvocab:
                resultwb.cell(row=n, column=j + 2, value=key)
                resultwb.cell(row=n, column=j + 3, value=value)
                n+=1


    file.save('E:\\桌面\\期货\\期货词频统计\\期货词频统计.xlsx')

shaixuanchanp()