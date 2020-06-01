import openpyxl
from utils import graph_tool
import datetime


file_path='E:\\桌面\\期货\\期货词频统计\\统计结果20200427.xlsx'
entity_path='E:\\桌面\\期货\\期货词频统计\\entity_20200526.xlsx'
relation_path='E:\\桌面\\期货\\期货词频统计\\relation_20200526.xlsx'
#file=openpyxl.load_workbook(file_path)



def get_entity():
    conceptobject=graph_tool.Concept()
    id_concept_dict=conceptobject.get_id_concept_dict()
    print(id_concept_dict)
    Date = datetime.datetime.now().date().strftime('%Y%m%d')
    #print(Date)
    file=graph_tool.TripleExtra()
    wsnames=file.getSheetName()
    allentity={}
    for ws in wsnames:
        wb=file.getwb(ws)
        if ws =='共有词':
            allentity=file.get_entity_id_label_list(wb,1,id_concept_dict,'期货相关词',allentity)
            allentity = file.get_entity_id_label_list(wb, 5, id_concept_dict, '能源期货相关词', allentity)
            allentity = file.get_entity_id_label_list(wb, 7, id_concept_dict, '农产品期货相关词', allentity)
            allentity = file.get_entity_id_label_list(wb, 9, id_concept_dict, '金属期货相关词', allentity)
        if ws =='能源期货相关词(按词性)':
            allentity['能源']=[]
            num = 0
            for j in range(3,wb.max_column+1,2):
                allentity = file.get_entity_id_label_list(wb, j, id_concept_dict, '', allentity)
                proitem={}

                proitem['id']=f"{id_concept_dict['能源']}_{Date}_{num}"
                num+=1
                proitem['concept']='能源'
                proitem['label']='期货:能源期货:能源'
                proitem['name']=wb.cell(1,j).value.split('期货')[0]
                allentity['能源'].append(proitem)
        if ws =='农产品期货相关词(按词性)':
            allentity['农产品']=[]
            num = 0
            for j in range(3,wb.max_column+1,2):
                allentity = file.get_entity_id_label_list(wb, j, id_concept_dict, '', allentity)
                proitem={}

                proitem['id']=f"{id_concept_dict['农产品']}_{Date}_{num}"
                num+=1
                proitem['concept']='农产品'
                proitem['label']='期货:农产品期货:农产品'
                proitem['name']=wb.cell(1,j).value.split('期货')[0]
                allentity['农产品'].append(proitem)
        if ws =='金属期货相关词(按词性)':
            allentity['金属']=[]
            num = 0
            for j in range(3,wb.max_column+1,2):
                allentity = file.get_entity_id_label_list(wb, j, id_concept_dict, '', allentity)
                proitem={}

                proitem['id']=f"{id_concept_dict['金属']}_{Date}_{num}"
                num+=1
                proitem['concept']='金属'
                proitem['label']='期货:金属期货:金属'
                proitem['name']=wb.cell(1,j).value.split('期货')[0]
                allentity['金属'].append(proitem)

    file.write_entity_to_excel(entity_path,allentity)

def get_relation():
    conceptobject = graph_tool.Concept()
    id_concept_dict = conceptobject.get_id_concept_dict()
    file=graph_tool.TripleExtra()
    wsnames = file.getSheetName()
    allrelation={}
    allrelation['子概念关系']=[]
    rel1={'head':'期货','headlabel':'期货','tail':'农产品期货','taillabel':'期货','出处':'交易手册'}
    rel2 = {'head': '期货', 'headlabel': '期货', 'tail': '能源期货', 'taillabel': '期货', '出处': '交易手册'}
    rel3 = {'head': '期货', 'headlabel': '期货', 'tail': '金属期货', 'taillabel': '期货', '出处': '交易手册'}
    allrelation['子概念关系'].append(rel1)
    allrelation['子概念关系'].append(rel2)
    allrelation['子概念关系'].append(rel3)
    for ws in wsnames:
        wb = file.getwb(ws)
        if ws == '共有词':
            allrelation=file.get_relation_inf(wb,1,'期货','期货','相关关系','交易手册',allrelation)
            allrelation =file.get_relation_inf(wb, 5, '能源期货', '期货', '相关关系', '交易手册', allrelation)
            allrelation = file.get_relation_inf(wb, 7, '农产品期货', '期货', '相关关系', '交易手册', allrelation)
            allrelation = file.get_relation_inf(wb, 9, '金属期货', '期货', '相关关系', '交易手册', allrelation)
        if ws =='能源期货相关词(按词性)':
            for j in range(3,wb.max_column+1,2):
                prod=wb.cell(1,j).value.split('期货')[0]
                prorel={'head': '能源期货', 'headlabel': '期货', 'tail': prod, 'taillabel': '能源', '出处': '交易手册'}
                allrelation['子概念关系'].append(prorel)
                allrelation = file.get_relation_inf(wb, j, prod, '能源', '相关关系', '交易手册', allrelation)
        if ws =='农产品期货相关词(按词性)':
            for j in range(3,wb.max_column+1,2):
                prod=wb.cell(1,j).value.split('期货')[0]
                prorel={'head': '农产品期货', 'headlabel': '期货', 'tail': prod, 'taillabel': '农产品', '出处': '交易手册'}
                allrelation['子概念关系'].append(prorel)
                allrelation = file.get_relation_inf(wb, j, prod, '农产品', '相关关系', '交易手册', allrelation)
        if ws =='金属期货相关词(按词性)':
            for j in range(3,wb.max_column+1,2):
                prod=wb.cell(1,j).value.split('期货')[0]
                prorel={'head': '金属期货', 'headlabel': '期货', 'tail': prod, 'taillabel': '金属', '出处': '交易手册'}
                allrelation['子概念关系'].append(prorel)
                allrelation = file.get_relation_inf(wb, j, prod, '金属', '相关关系', '交易手册', allrelation)


    file.write_relation_to_excel(relation_path, allrelation)

get_entity()