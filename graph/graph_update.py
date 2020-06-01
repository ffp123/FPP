from py2neo import Graph
from ast import literal_eval
import json
from utils import graph_tool
import openpyxl

concept_path='E:\\桌面\\期货\\Concept_1.0.xlsx'
entity_path='E:\\桌面\\期货\\期货词频统计\\entity_20200427.xlsx'
relation_path='E:\\桌面\\期货\\期货词频统计\\relation_20200427.xlsx'

graphobject=graph_tool.graphutil('localhost')
graph=graphobject.graph()

#本体层更新
def update_concept():
    graphobject.clear(graph,'期货相关词图谱')
    concept_file=graph_tool.Concept()
    sheets=concept_file.getSheetName()
    for ws in sheets:
        wb=concept_file.getwb(ws)
        if ws=='节点':
            for i in range(1,wb.max_row+1):
                if wb.cell(i,1).value != None and wb.cell(i,1).value != '' and wb.cell(i,1).value != 'None':
                    attr = {'id': wb.cell(i,1).value, 'name': wb.cell(i,2).value}
                    if wb.cell(i,3).value!=None and wb.cell(i,3).value != '' and wb.cell(i,3).value != 'None':
                        for attr_ in wb.cell(i,3).value.split('|'):
                            attr[attr_]=''
                    graphobject.createNodeWithAttribute(graph,':'.join(wb.cell(i,4).value.split('|')), attr)

        if ws=='关系':
            for i in range(1,wb.max_row+1):
                if wb.cell(i,1).value != None and wb.cell(i,1).value != '' and wb.cell(i,1).value != 'None':
                    attr = {'name': wb.cell(i,3).value}
                    if wb.cell(i,4).value != None and wb.cell(i,4).value != '' and wb.cell(i,4).value != 'None':
                        for attr_ in wb.cell(i,4).value.split('|'):
                            attr[attr_]=''
                    graphobject.createRelationshipByNameWithAttribute(graph,wb.cell(i,2).value,wb.cell(i,1).value,wb.cell(i,6).value,wb.cell(i,5).value,wb.cell(i,3).value,attr)


#更新实体和关系
def update_entity_relation():
    entity_file=openpyxl.load_workbook(entity_path)
    relation_file=openpyxl.load_workbook(relation_path)
    entity_sheets=entity_file.sheetnames
    relation_sheets=relation_file.sheetnames

    for enws in entity_sheets:
        enwb=entity_file[enws]
        for i in range(1,enwb.max_row+1):
            if enwb.cell(i, 1).value != None and enwb.cell(i, 1).value != '' and enwb.cell(i, 1).value != 'None':
                if graphobject.existNode(graph,enws,enwb.cell(i,2).value):
                    graphobject.addNodeLabel(graph,enws,enwb.cell(i,2).value,enwb.cell(i,3).value)
                else:
                    label=enwb.cell(i,3).value
                    attr = {'id': enwb.cell(i,1).value, 'name': enwb.cell(i,2).value, 'concept': enwb.cell(i,4).value}
                    graphobject.createNodeWithAttribute(graph,label, attr)
    for rews in relation_sheets:
        rewb=relation_file[rews]
        for i in range(1,rewb.max_row+1):
            if rewb.cell(i, 1).value != None and rewb.cell(i, 1).value != '' and rewb.cell(i, 1).value != 'None':
                if graphobject.existTriple(graph,rewb.cell(i, 2).value,rewb.cell(i, 1).value,rewb.cell(i, 4).value,rewb.cell(i, 3).value,rews):
                    triple=graphobject.getSingleTriple(graph,rewb.cell(i, 2).value,rewb.cell(i, 1).value,rewb.cell(i, 4).value,rewb.cell(i, 3).value,rews)[0]
                    cclist=[]
                    cclist.append(triple[2]['出处'])
                    cclist.append(rewb.cell(i, 5).value)
                    cc = list(set(cclist))
                    graphobject.updateTripleAttribute(graph,rewb.cell(i, 2).value,rewb.cell(i, 1).value,rewb.cell(i, 4).value,rewb.cell(i, 3).value,rews,{'出处':str(cc)})
                    if rews in ['相关关系']:
                        graphobject.updateTripleAttribute(graph, rewb.cell(i, 2).value, rewb.cell(i, 1).value,
                                                         rewb.cell(i, 4).value, rewb.cell(i, 3).value, rews,
                                                         {'相关性':str(rewb.cell(i, 6).value)})
                else:
                    attr = {'name': rews, '出处': rewb.cell(i, 5).value}
                    if rews in ['相关关系']:
                        attr['相关性'] = rewb.cell(i, 6).value
                    graphobject.createRelationshipByNameWithAttribute(graph,rewb.cell(i, 1).value,rewb.cell(i, 2).value,rewb.cell(i, 3).value,rewb.cell(i, 4).value,rews,attr)

update_entity_relation()