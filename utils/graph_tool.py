from py2neo import Graph
from ast import literal_eval
import json
import openpyxl
import datetime

cix_label={'n':'其他相关词','ng':'其他相关词','nr':'人物相关词','nrfg':'人物相关词','nrt':'人物相关词',
              'ns':'地域相关词','nt':'人物相关词','nz':'其他相关词','s':'地域相关词','t':'时间相关词','tg':'时间相关词','vn':'其他相关词'}

class Concept(object):
    def __init__(self, concept_path = 'E:\\桌面\\期货\\Concept_1.0.xlsx'):
        self.concept_file = openpyxl.load_workbook(concept_path)

    def getSheetName(self):
        return self.concept_file.sheetnames

    def getwb(self,ws):
        return self.concept_file[ws]

    def get_id_concept_dict(self):
        nodews=self.concept_file['节点']
        id_concept_dict={}
        for i in range(1,nodews.max_row+1):
            if nodews.cell(i,2).value!=None and nodews.cell(i,2).value!='':
                id_concept_dict[nodews.cell(i,2).value]=nodews.cell(i,1).value
        return id_concept_dict

class TripleExtra(object):

    def __init__(self, file_path='E:\\桌面\\期货\\期货词频统计\\统计结果20200427.xlsx'):
        self.file = openpyxl.load_workbook(file_path)

    def getSheetName(self):
        return self.file.sheetnames

    def getwb(self,ws):
        return self.file[ws]

    def get_entity_id_label_list(self,wb,j,iddict,label,allentity):
        Date = datetime.datetime.now().date().strftime('%Y%m%d')

        for i in range(2,wb.max_row+1):
            if wb.cell(i,j).value!=None and wb.cell(i,j).value!='':
                cix=wb.cell(i,j).value.split('|')[1]
                concept=cix_label[cix]
                id=f'{iddict[concept]}_{Date}'
                itemdict = {}
                itemdict['concept'] = concept
                itemdict['name'] = wb.cell(i, j).value.split('|')[0]
                if label!='':
                    itemdict['label'] = f'{cix_label[cix]}:{label}'
                else:
                    itemdict['label'] = f'{cix_label[cix]}'
                if concept not in allentity:
                    allentity[concept]=[]
                    num=0
                    itemdict['id'] = f'{id}_{num}'
                    allentity[concept].append(itemdict)
                else:
                    num=len(allentity[concept])
                    itemdict['id'] = f'{id}_{num}'
                    allentity[concept].append(itemdict)

        return allentity

    def get_relation_inf(self,wb,j,head,headlabel,relationtype,chuchu,allrelation):
        if j==1:
            for i in range(2,wb.max_row+1):
                if wb.cell(i,j).value!=None and wb.cell(i,j).value!='':
                    word=wb.cell(i,j).value.split('|')[0]
                    cix=wb.cell(i,j).value.split('|')[1]
                    if relationtype not in allrelation:
                        allrelation[relationtype]=[]
                        onerel={'head':head,'headlabel':headlabel,'tail':word,'taillabel':cix_label[cix],'出处':chuchu}
                        if relationtype=='相关关系':
                            onerel['相关性']=f'能源_{wb.cell(i,j+1).value}_农产品_{wb.cell(i,j+2).value}_金属_{wb.cell(i,j+3).value}'
                        allrelation[relationtype].append(onerel)
                    else:
                        onerel = {'head': head, 'headlabel': headlabel, 'tail': word, 'taillabel': cix_label[cix], '出处': chuchu}
                        if relationtype == '相关关系':
                            onerel['相关性'] = f'能源_{wb.cell(i,j+1).value}_农产品_{wb.cell(i,j+2).value}_金属_{wb.cell(i,j+3).value}'
                        allrelation[relationtype].append(onerel)
            return allrelation
        else:
            for i in range(2,wb.max_row+1):
                if wb.cell(i, j).value != None and wb.cell(i, j).value != '':
                    word=wb.cell(i,j).value.split('|')[0]
                    cix=wb.cell(i,j).value.split('|')[1]
                    if relationtype not in allrelation:
                        allrelation[relationtype]=[]
                        onerel={'head':head,'headlabel':headlabel,'tail':word,'taillabel':cix_label[cix],'出处':chuchu}
                        if relationtype=='相关关系':
                            onerel['相关性']=wb.cell(i,j+1).value
                        allrelation[relationtype].append(onerel)
                    else:
                        onerel = {'head': head, 'headlabel': headlabel, 'tail': word, 'taillabel': cix_label[cix], '出处': chuchu}
                        if relationtype == '相关关系':
                            onerel['相关性'] = wb.cell(i,j+1).value
                        allrelation[relationtype].append(onerel)
            return allrelation



    def write_entity_to_excel(self,filepath,allentity):
        entity_file=openpyxl.Workbook()
        for type,entities in allentity.items():
            wb=entity_file.create_sheet(type)
            i=1
            for entity in entities:
                wb.cell(row=i, column=1, value=entity['id'])
                wb.cell(row=i,column=2,value=entity['name'])
                wb.cell(row=i, column=3, value=entity['label'])
                wb.cell(row=i, column=4, value=entity['concept'])
                i+=1
        entity_file.save(filepath)

    def write_relation_to_excel(self,filepath,allrelation):
        relation_file=openpyxl.Workbook()
        for reltype,relations in allrelation.items():
            wb=relation_file.create_sheet(reltype)
            i=1
            for relation in relations:
                wb.cell(row=i, column=1, value=relation['head'])
                wb.cell(row=i, column=2, value=relation['headlabel'])
                wb.cell(row=i, column=3, value=relation['tail'])
                wb.cell(row=i, column=4, value=relation['taillabel'])
                wb.cell(row=i, column=5, value=relation['出处'])
                if reltype in ['相关关系']:
                    wb.cell(row=i, column=6, value=relation['相关性'])
                i+=1
        relation_file.save(filepath)

class graphutil():
    def __init__(self, host):

        self.__graph = Graph(f"http://{host}:7474", username="neo4j", password="taohu0311")

    def graph(self):
        return self.__graph

    def clear(self,graph,label):
        cql = f'match (h:{label})-[r]-(t) delete r, h'
        print(cql)
        graph.run(cql)
        cql = f'match (h:{label}) delete h'
        print(cql)
        graph.run(cql)

    def createNodeWithAttribute(self,graph,label,attribute):
        attr = '{'
        for k, v in attribute.items():
            attr += f'{k}: "{v}",'
        attr = attr[:-1] + '}'
        cql = f'merge (:{label}  {attr})'
        print(cql)
        graph.run(cql)

    def createRelationshipByNameWithAttribute(self,graph,h_name, h_label, t_name, t_label, r_name, attribute):
        attr = '{'
        for k, v in attribute.items():
            attr += f'{k}: "{v}",'
        attr = attr[:-1] + '}'
        cql = f'match (h:{h_label}),(t:{t_label}) where h.name="{h_name}" and t.name="{t_name}" ' \
                f'merge (h)-[r:{r_name} {attr}]->(t)'
        print(cql)
        graph.run(cql)

    def existNode(self,graph, bt, name):
        cql = f'match (n:{bt}) where n.name="{name}" return count(*) as n'
        print(cql)
        n = graph.run(cql)
        return list(n)[0]['n'] == 1

    def addNodeLabel(self,graph, bt, name, label):
        cql = f'match (n:{bt}) where n.name="{name}" set n:{label}'
        print(cql)
        graph.run(cql)

    def existTriple(self,graph,h_label, h_name, t_label, t_name, r_name):
        cql = f'match (h:{h_label})-[r:{r_name}]->(t:{t_label}) where h.name="{h_name}" and t.name="{t_name}" return count(*) as n'
        print(cql)
        n = graph.run(cql)
        return list(n)[0]['n'] == 1

    def getSingleTriple(self,graph, h_label, h_name, t_label, t_name, r_name):
        cql = f'match (h:{h_label})-[r:{r_name}]->(t:{t_label}) where h.name="{h_name}" and t.name="{t_name}" return r'
        print(cql)
        ts = graph.run(cql)
        triples = []
        for t in ts:
            triples.append((h_label, h_name, {k: v for k, v in t['r'].items()}, t_label, t_name))
        return triples
    def updateTripleAttribute(self,graph, h_label, h_name, t_label, t_name,r_name, attribute):
        for k, v in attribute.items():
            cql = f'match (h:{h_label} )-[r:{r_name}]->(t:{t_label}) where h.name="{h_name}" and t.name="{t_name}" ' \
                    f'set r.{k}="{v}"'
            print(cql)
            graph.run(cql)
