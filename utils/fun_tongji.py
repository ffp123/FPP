import re
import openpyxl
import ast

num_ncp=15
num_ny=17
num_js=5

class Excel(object):

    def __init__(self, file_path):
        self.file = openpyxl.load_workbook(file_path)

    def getSheetName(self):
        return self.file.sheetnames

    def getwb(self,ws):
        return self.file[ws]

    def createsheet(self,ws):
        return self.file.create_sheet(ws)

    def save(self,path):
        self.file.save(path)


class tongji(object):

    # 对字典中的元素按值降序排列
    def paixu(self,vocab):
        sortvocab = sorted(vocab.items(), key=lambda item: item[1], reverse=True)
        sortdict = {}
        for item in sortvocab:
            # print(item[0], item[1])
            sortdict[item[0]] = item[1]

        return sortdict
    # 统计每个期货词频
    def tongjichanpin(self,worksheets,file1,resultwb):
        allvocab = {}
        print(worksheets)
        j = 1
        for ws in worksheets:
            if ws != '交易手册划分汇总':
                wb2 = file1.getwb(ws)
                itemvocab = {}

                for i in range(2, wb2.max_row + 1):
                    if wb2.cell(i, 1).value != None and wb2.cell(i, 1).value != '':
                        worddict = ast.literal_eval(wb2.cell(i, 4).value)
                        # print(worddict)
                        # print(type(worddict))
                        for k, v in worddict.items():
                            if k not in itemvocab:
                                itemvocab[k] = v
                            else:
                                itemvocab[k] += v
                            if k not in allvocab:
                                allvocab[k] = v
                            else:
                                allvocab[k] += v
                sortdict = self.paixu(itemvocab)
                resultwb.cell(row=1, column=j, value=str(ws + '期货词'))
                resultwb.cell(row=1, column=j + 1, value='词频统计')
                n = 2
                for key, value in sortdict.items():
                    resultwb.cell(row=n, column=j, value=key)
                    resultwb.cell(row=n, column=j + 1, value=value)
                    n += 1
                j += 2
        allsortdict = self.paixu(allvocab)
        return allsortdict

    # 统计所有期货词频
    def tonngjiall(self,dirs,path):
        allvocab = {}
        for dir in dirs:
            if dir != '词频统计.xlsx':
                file = openpyxl.load_workbook(path + '\\' + dir)
                worksheets = file.sheetnames
                for ws in worksheets:
                    if ws != '交易手册划分汇总':
                        wb = file[ws]
                        for i in range(2, wb.max_row + 1):
                            if wb.cell(i, 1).value != None and wb.cell(i, 1).value != '':
                                worddict = ast.literal_eval(wb.cell(i, 4).value)
                                # print(worddict)
                                # print(type(worddict))
                                for k, v in worddict.items():
                                    if k not in allvocab:
                                        allvocab[k] = v
                                    else:
                                        allvocab[k] += v

        sortdict = self.paixu(allvocab)
        return sortdict

    # 从工作表中获取指定列的词以及词频字典
    def get_word_freq_dict(self,ws, j):
        worddict = {}
        wordvocab = []
        for i in range(2, ws.max_row + 1):
            word = ws.cell(i, j).value
            if word != '' and word != None:
                cipin = int(ws.cell(i, j + 1).value)
                worddict[word] = cipin
                wordvocab.append(word)
        return worddict, wordvocab

    # 计算置信度
    def calzhixindu(self,ncpfreq, nyfreq, jsfreq):
        allfreq = ncpfreq + nyfreq + jsfreq
        num_all = num_ncp + num_js + num_ny
        mean = float(allfreq / num_all)
        mean = float(mean / 3)
        ncppercent = float(ncpfreq / allfreq)
        nypercent = float(nyfreq / allfreq)
        jspercent = float(jsfreq / allfreq)
        ncpflag = 1
        nyflag = 1
        jsflag = 1
        ncpscore = float(ncpfreq / num_ncp)
        nyscore = float(nyfreq / num_ny)
        jsscore = float(jsfreq / num_js)
        if float(ncppercent * ncpscore) < float(mean * 0.6):
            ncpflag = 0
        if float(nypercent * nyscore) < float(mean * 0.6):
            nyflag = 0
        if float(jspercent * jsscore) < float(mean * 0.6):
            jsflag = 0

        return ncpflag, nyflag, jsflag

    # 统计各类期货专有词与共有词
    def tongjicom_and_spec(self,ws):
        ncpdict, ncpvocab = self.get_word_freq_dict(ws, 5)
        jsdict, jsvocab = self.get_word_freq_dict(ws, 7)
        nydict, nyvocab = self.get_word_freq_dict(ws, 3)
        jj1 = list(set(ncpvocab).intersection(set(nyvocab)))
        jj2 = list(set(jj1).intersection(set(jsvocab)))
        ncp_jj2 = list(set(ncpvocab) - set(jj2))
        ny_jj2 = list(set(nyvocab) - set(jj2))
        js_jj2 = list(set(jsvocab) - set(jj2))
        for item in jj2:
            ncpflag, nyflag, jsflag = self.calzhixindu(ncpdict[item], nydict[item], jsdict[item])
            if ncpflag == 1 and nyflag == 1 and jsflag == 1:
                continue
            else:
                jj2.remove(item)
                ncp_jj2.append(item)
                ny_jj2.append(item)
                js_jj2.append(item)
        jj2dict = {}
        for item in jj2:
            jj2dict[item] = nydict[item]
        sortjj2dict = self.paixu(jj2dict)
        ncpjj2dict = {}
        for item in ncp_jj2:
            ncpjj2dict[item] = ncpdict[item]
        sortncpjj2dict = self.paixu(ncpjj2dict)
        nyjj2dict = {}
        for item in ny_jj2:
            nyjj2dict[item] = nydict[item]
        sortnyjj2dict = self.paixu(nyjj2dict)
        jsjj2dict = {}
        for item in js_jj2:
            jsjj2dict[item] = jsdict[item]
        sortjsjj2dict = self.paixu(jsjj2dict)

        return sortjj2dict,sortncpjj2dict,sortnyjj2dict,sortjsjj2dict,ncpdict,nydict,jsdict

    # 统计每种期货专有词
    def shaixuanchanp(self,resultwb,prowb):

        commonvocab = []
        for i in range(2, resultwb.max_row + 1):
            if resultwb.cell(i, 1).value != None and resultwb.cell(i, 1).value != '':
                commonvocab.append(resultwb.cell(i, 1).value)

        for j in range(1, prowb.max_column + 1, 2):
            resultwb.cell(row=1, column=j + 2, value=str(prowb.cell(1, j).value + '(专有)'))
            resultwb.cell(row=1, column=j + 3, value=prowb.cell(1, j + 1).value)
            prodict = {}
            provocab = []
            for i in range(2, prowb.max_row + 1):
                if prowb.cell(i, j).value != None and prowb.cell(i, j).value != '':
                    provocab.append(prowb.cell(i, j).value)
                    prodict[prowb.cell(i, j).value] = prowb.cell(i, j + 1).value
            n = 2
            for key, value in prodict.items():
                if key in provocab and key not in commonvocab:
                    resultwb.cell(row=n, column=j + 2, value=key)
                    resultwb.cell(row=n, column=j + 3, value=value)
                    n += 1

    def get_all_word_freq(self,ws):
        all_word_freq = {}
        for j in range(3, ws.max_column + 1, 2):
            itemdict, itemvocab = self.get_word_freq_dict(ws, j)
            all_word_freq[ws.cell(1, j).value] = itemdict

        return all_word_freq

    #获取各类期货下的共有词
    def extract(self,allwordws,productws):
        allword, allwordvocab = self.get_word_freq_dict(allwordws, 7)
        productword = self.get_all_word_freq(productws)
        word_in_common = []
        word_score_dict = {}
        pro_num = len(productword)
        for word, freq in allword.items():
            num = 0
            shaixuan = {}
            for prod, worddict in productword.items():
                if word in worddict:
                    num += 1
                    shaixuan[prod] = worddict[word]
            if num >= len(productword):
                # flag=fun_tongji.judge_word(shaixuan,pro_num)
                # if flag>=len(shaixuan):
                word_in_common.append(word)
                word_score_dict[word] = len(shaixuan) / pro_num
        return word_score_dict,productword

    # 获取各种词性下的词以及词频，并降序排列
    def fencen(self,vocabdict):
        label_dict = {}
        for key, value in vocabdict.items():
            label = get_alpha_str(key)
            word = key.replace(label, '')
            if label not in label_dict:
                label_dict[label] = {}
                label_dict[label][word] = value
            else:
                label_dict[label][word] = value

        for key, value in label_dict.items():
            sorteach = sorted(value.items(), key=lambda item: item[1], reverse=True)
            sorteachdict = {}
            for item in sorteach:
                sorteachdict[item[0]] = item[1]
            label_dict[key] = sorteachdict

        return label_dict
    # 按词性分类
    def fenlei(self,ws,resultws):

        for j in range(1, ws.max_column + 1, 2):
            # j=1
            nvdict = {}
            ncpdict = {}
            jsdict = {}
            for i in range(2, ws.max_row + 1):
                if ws.cell(i, j).value != '' and ws.cell(i, j).value != None:
                    # print(type(ws.cell(i,j+1).value))
                    nvdict[ws.cell(i, j).value] = ws.cell(i, j + 1).value
                    # nvdict[ws.cell(i,1).value]=ws.cell(i,2).value
                    # ncpdict[ws.cell(i, 1).value] = int(ws.cell(i, 3).value)
                    # jsdict[ws.cell(i, 1).value] = int(ws.cell(i, 4).value)

            labeldict = self.fencen(nvdict)
            print(labeldict)
            recordrow = 2
            for label, wordfre in labeldict.items():
                for word, freq in wordfre.items():
                    resultws.cell(row=recordrow, column=j, value=str(word + '|' + label))
                    resultws.cell(row=recordrow, column=j + 1, value=nvdict[str(word + label)])
                    # resultws.cell(row=recordrow,column=1,value=str(word+'|'+label))
                    # resultws.cell(row=recordrow, column=2, value=nvdict[str(word+label)])
                    # resultws.cell(row=recordrow, column=3, value=ncpdict[str(word+label)])
                    # resultws.cell(row=recordrow, column=4, value=jsdict[str(word+label)])
                    recordrow += 1


#获取字符串中的字母，用来获取词性
def get_alpha_str(s):
    result = re.sub(r'[^a-z]', '', s)
    return result


def judge_word(shaixuan,num):
    allfreq=0
    flag=0
    for key,value in shaixuan.items():
        allfreq+=value
    mean=allfreq/num
    for key,value in shaixuan.items():
        if value>=mean:
            flag+=1
    return flag